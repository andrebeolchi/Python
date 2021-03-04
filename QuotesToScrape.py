import pandas as pd
from bs4 import BeautifulSoup
from urllib.request import urlopen 
import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
import xlsxwriter
import matplotlib.pyplot as plt
import numpy as np
import PyQt5
import requests
from collections import Counter

wbCriado = openpyxl.Workbook('QuotesToScrape.xlsx')

def msgIniciando(): # SOMENTE PERFUMARIA

    print("\033[1;32m╔═══════════════════════════════════════════════════════════╗")
    print("\033[1;32m║    \033[1;33mO Sistema está iniciando, Aguarde alguns instantes!    \033[1;32m║")
    print("\033[1;32m╚═══════════════════════════════════════════════════════════╝")

def msgConcluido(): # SOMENTE PERFUMARIA

    print("\033[1;35m╔═══════════════════════════════════════════════════════════╗")
    print("\033[1;35m║              \033[1;33mArquivo Exportado com Sucesso!!              \033[1;35m║")
    print("\033[1;35m╚═══════════════════════════════════════════════════════════╝")

def msgAguarde(): # SOMENTE PERFUMARIA
    if x < 10:
        print("\033[1;34m╠═══════════════════════════════════════════════════════════╣")
        print("\033[1;34m║                  \033[1;33mPegando dados da Página 0" + str(x) +"               \033[1;34m║")
        print("\033[1;34m╠═══════════════════════════════════════════════════════════╣")
    else:
        print("\033[1;34m╠═══════════════════════════════════════════════════════════╣")
        print("\033[1;34m║                  \033[1;33mPegando dados da Página " + str(x) +"               \033[1;34m║")
        print("\033[1;34m╠═══════════════════════════════════════════════════════════╣")

def msgGraf():# SOMENTE PERFUMARIA
    print("\033[1;34m╠═══════════════════════════════════════════════════════════╣")
    print("\033[1;34m║              \033[1;33mCriando e Exportando o Gráfico               \033[1;34m║")
    print("\033[1;34m╚═══════════════════════════════════════════════════════════╝")

listaFrase = []
listaAutor = []
listaAutor2 = []
listaFrase2 = []
maxPag = 10 # Quantidade de páginas do site

msgIniciando()

df = pd.DataFrame(columns=['Frase','Autor']) # Criando DataFrame das Frases

for x in range(maxPag):
    x += 1
    
    url = "http://quotes.toscrape.com/page/" + str(x) # Definindo o site

    html = urlopen(url)

    bs = BeautifulSoup(html, 'lxml')

    textos = bs.select('div', class_="row")

    frase = bs.find_all('span', class_='text') # CAPTURANDO FRASES
    autor = bs.find_all('small', class_='author') # CAPTURANDO AUTOR

    listaFrase = [x.text for x in frase] # CAPTURANDO SÓ A FRASE
    listaAutor = [y.text for y in autor] # CAPTURANDO SÓ O AUTOR

    listaAutor2.extend(listaAutor) # Juntando os autores de todas as páginas
    listaFrase2.extend(listaFrase) # Juntando as frases de todas as páginas

    msgAguarde()

df['Frase'] = listaFrase2 # Escrevendo na coluna Frase 
df['Autor'] = listaAutor2 # Escrevendo na coluna Autor

qtdAutor = []

for x in range(maxPag): # Verificando quantas vezes o autor aparece no site
    qtdAutor = dict((x,listaAutor2.count(x)) for x in set(listaAutor2))
    lst = [i for i in qtdAutor if isinstance(i, (int, float))]
    x += 1

dn = pd.DataFrame(columns=['Autor','Quantidade']) # DataFrame para Autor/Quantidade

keys = qtdAutor.keys() # Separando Autor
values = qtdAutor.values() # Separando Quantidade

dn['Autor'] = keys # Escrevendo na coluna Autor
dn['Quantidade'] = values # Escrevendo na coluna Quantidade



book = load_workbook('QuotesToScrape.xlsx')


with pd.ExcelWriter('QuotesToScrape.xlsx', engine='openpyxl') as writer: # pylint: disable=abstract-class-instantiated  # Exportando os DataFrames para o mesmo arquivo Excel, Comentário anterior para não apresentar erro do pylint
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, "Frases", columns=['Frase', 'Autor'])
    dn.to_excel(writer, "Gráfico", columns=['Autor', 'Quantidade'])

df.to_csv('testeFRASE.csv') # Exportando Frases para .CSV
dn.to_csv('testeQTD.csv') # Exportando Quantidades para .CSV


###############################################
##########     Criando o gráfico     ##########
###############################################
msgGraf()

dn2 = dn.sort_values('Quantidade') # Sorteando a tabela de quantidade de F/A
qtd = dn2['Quantidade'] # Pegando a quantidade da nova tabela
aut = dn2['Autor'] # Pegando a pegando o autor da nova tabela

fig, ax = plt.subplots(figsize=(15, 15)) # Criando o gráfico com tamanho 15x15

ax.barh(aut, qtd, align='center', color='blue') # Criando gráfico e atribuindo o Autor, a Quantidade e alinhando no centro
ax.invert_yaxis()  # Inverte o gráfico (de baixo para cima)
ax.set_xlabel('Quantidade') # Legenda do eixo x
ax.set_title('Quantidade de Frases por Autor') # Nome do Gráfico
for i, qtd in enumerate(qtd): # Pegando/Apresentando os valores individuais depois de cada barra do gráfico
    ax.text(qtd + 0.15, i + .25, str(qtd), color='blue', fontweight='bold')

plt.savefig('grafico.png', dpi = 200) # Salvando a imagem (grafico.png)

#plt.show() # Para ver o gráfico aqui

#################################################
## Exportando a imagem do gráfico para o Excel ##
#################################################

wb = openpyxl.load_workbook('QuotesToScrape.xlsx') # Abrindo o arquivo excel
ws = wb.worksheets[2] # Abrindo a 3ª planilha do arquivo

img = openpyxl.drawing.image.Image('grafico.png') # Criando a variavel com a imagem

ws.add_image(img) # Adicionando a imagem na planilha

wb.save('QuotesToScrapeG.xlsx') # Salvando o arquivo excel e substitui o nome

msgConcluido() # Perfumaria